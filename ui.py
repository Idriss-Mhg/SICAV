import os
import subprocess
import sys
import threading
import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext, ttk

import openpyxl
from docx import Document

from main import run


# ─────────────────────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────────────────────

def _browse(var, filetypes, save=False):
    if save:
        path = filedialog.asksaveasfilename(
            filetypes=filetypes, defaultextension=filetypes[0][1])
    else:
        path = filedialog.askopenfilename(filetypes=filetypes)
    if path:
        var.set(path)


def _open_file(path):
    if sys.platform.startswith("win"):
        os.startfile(path)
    elif sys.platform == "darwin":
        subprocess.call(["open", path])
    else:
        subprocess.call(["xdg-open", path])


# ─────────────────────────────────────────────────────────────────────────────
# Anchor picker dialog
# ─────────────────────────────────────────────────────────────────────────────

class AnchorPickerDialog(tk.Toplevel):
    """
    Scans a .docx and shows all unique non-empty paragraphs in a searchable
    listbox. The selected text is written into `target_var`.
    """

    def __init__(self, parent, docx_path, target_var):
        super().__init__(parent)
        self.title("Choisir une ancre")
        self.transient(parent)
        self.grab_set()
        self.resizable(True, True)
        self.minsize(560, 420)
        self.target_var = target_var
        self.all_paragraphs = []
        self._build(docx_path)

    def _build(self, docx_path):
        # ── Search bar ────────────────────────────────────────────────────────
        frm_top = ttk.Frame(self, padding=(10, 8, 10, 4))
        frm_top.pack(fill="x")
        ttk.Label(frm_top, text="Filtrer :").pack(side="left")
        self.var_search = tk.StringVar()
        self.var_search.trace_add("write", lambda *_: self._filter())
        ttk.Entry(frm_top, textvariable=self.var_search, width=50).pack(
            side="left", padx=(6, 0), fill="x", expand=True)

        # ── Listbox ───────────────────────────────────────────────────────────
        frm_list = ttk.Frame(self, padding=(10, 0, 10, 4))
        frm_list.pack(fill="both", expand=True)

        self.listbox = tk.Listbox(
            frm_list, selectmode="single",
            font=("Courier New", 9), activestyle="dotbox",
        )
        sb = ttk.Scrollbar(frm_list, orient="vertical",
                           command=self.listbox.yview)
        self.listbox.config(yscrollcommand=sb.set)
        self.listbox.pack(side="left", fill="both", expand=True)
        sb.pack(side="right", fill="y")
        self.listbox.bind("<Double-Button-1>", lambda _: self._select())

        # ── Buttons ───────────────────────────────────────────────────────────
        frm_btn = ttk.Frame(self, padding=(10, 4, 10, 10))
        frm_btn.pack(fill="x")
        ttk.Button(frm_btn, text="Sélectionner",
                   command=self._select).pack(side="right")
        ttk.Button(frm_btn, text="Annuler",
                   command=self.destroy).pack(side="right", padx=(0, 6))

        # ── Load paragraphs in background ─────────────────────────────────────
        self.listbox.insert("end", "Chargement du document…")
        self.listbox.config(state="disabled")
        threading.Thread(target=self._load, args=(docx_path,),
                         daemon=True).start()

    def _load(self, docx_path):
        try:
            doc = Document(docx_path)
            seen = set()
            paras = []
            for p in doc.paragraphs:
                t = p.text.strip()
                if t and t not in seen:
                    seen.add(t)
                    paras.append(t)
            self.all_paragraphs = paras
            self.after(0, self._populate)
        except Exception as exc:
            self.after(0, lambda: messagebox.showerror(
                "Erreur", f"Impossible de lire le document :\n{exc}", parent=self))

    def _populate(self):
        self.listbox.config(state="normal")
        self.listbox.delete(0, "end")
        for t in self.all_paragraphs:
            self.listbox.insert("end", t)

    def _filter(self):
        query = self.var_search.get().lower()
        self.listbox.delete(0, "end")
        for t in self.all_paragraphs:
            if query in t.lower():
                self.listbox.insert("end", t)

    def _select(self):
        sel = self.listbox.curselection()
        if not sel:
            return
        self.target_var.set(self.listbox.get(sel[0]))
        self.destroy()


# ─────────────────────────────────────────────────────────────────────────────
# Tab 1 — Traitement
# ─────────────────────────────────────────────────────────────────────────────

class TabTraitement(ttk.Frame):
    def __init__(self, parent, shared):
        super().__init__(parent)
        self.shared = shared   # shared vars (docx path, excel path)
        self._build()

    def _build(self):
        pad = {"padx": 12, "pady": 6}

        # ── Files ─────────────────────────────────────────────────────────────
        frm_files = ttk.LabelFrame(self, text="Fichiers", padding=10)
        frm_files.pack(fill="x", **pad)
        frm_files.columnconfigure(1, weight=1)

        self.var_out    = tk.StringVar()
        self.var_review = tk.StringVar()
        self.var_pdf    = tk.StringVar()

        self._file_row(frm_files, 0, "Prospectus (.docx)",
                       self.shared["var_docx"],
                       [("Word documents", "*.docx")])
        self._file_row(frm_files, 1, "Mapping Excel (.xlsx)",
                       self.shared["var_excel"],
                       [("Excel files", "*.xlsx")])
        self._file_row(frm_files, 2, "Sortie normale (.docx)",
                       self.var_out,
                       [("Word documents", "*.docx")], save=True)
        self._file_row(frm_files, 3, "Sortie review (.docx)",
                       self.var_review,
                       [("Word documents", "*.docx")], save=True)
        self._file_row(frm_files, 4, "Sortie PDF (.pdf)",
                       self.var_pdf,
                       [("PDF", "*.pdf")], save=True)

        # ── Options ───────────────────────────────────────────────────────────
        frm_opts = ttk.LabelFrame(self, text="Options", padding=10)
        frm_opts.pack(fill="x", **pad)

        self.var_open_after = tk.BooleanVar(value=True)
        self.var_dry_run    = tk.BooleanVar(value=False)

        ttk.Checkbutton(frm_opts,
                        text="Ouvrir le fichier de sortie après traitement",
                        variable=self.var_open_after).grid(row=0, column=0, sticky="w")
        ttk.Checkbutton(frm_opts,
                        text="Dry run  (affiche les insertions sans sauvegarder)",
                        variable=self.var_dry_run).grid(row=1, column=0, sticky="w")

        # ── Run ───────────────────────────────────────────────────────────────
        frm_run = ttk.Frame(self)
        frm_run.pack(fill="x", padx=12, pady=(4, 0))

        self.btn_run = ttk.Button(frm_run, text="▶  Lancer",
                                  command=self._on_run, width=18)
        self.btn_run.pack(side="right")
        ttk.Button(frm_run, text="Effacer le log",
                   command=self._clear_log).pack(side="right", padx=(0, 8))

        # ── Progress ──────────────────────────────────────────────────────────
        self.progress = ttk.Progressbar(self, mode="indeterminate")
        self.progress.pack(fill="x", padx=12, pady=(6, 0))

        # ── Log ───────────────────────────────────────────────────────────────
        frm_log = ttk.LabelFrame(self, text="Log", padding=6)
        frm_log.pack(fill="both", expand=True, padx=12, pady=8)

        self.log_box = scrolledtext.ScrolledText(
            frm_log, height=12, wrap="word", state="disabled",
            font=("Courier New", 9), bg="#1e1e1e", fg="#d4d4d4",
        )
        self.log_box.pack(fill="both", expand=True)
        self.log_box.tag_config("ok",   foreground="#4ec9b0")
        self.log_box.tag_config("warn", foreground="#f4c15a")
        self.log_box.tag_config("err",  foreground="#f44747")
        self.log_box.tag_config("head", foreground="#9cdcfe")

    def _file_row(self, parent, row, label, var, filetypes, save=False):
        ttk.Label(parent, text=label).grid(row=row, column=0, sticky="w", pady=3)
        ttk.Entry(parent, textvariable=var).grid(
            row=row, column=1, sticky="ew", padx=(8, 4))
        ttk.Button(parent, text="…", width=3,
                   command=lambda v=var, ft=filetypes, s=save: _browse(v, ft, s),
                   ).grid(row=row, column=2)

    def _log(self, msg, tag=None):
        def _w():
            self.log_box.config(state="normal")
            self.log_box.insert("end", msg + "\n", tag or "")
            self.log_box.see("end")
            self.log_box.config(state="disabled")
        self.after(0, _w)

    def _smart_log(self, msg):
        if msg.strip().startswith("✓"):
            self._log(msg, "ok")
        elif msg.strip().startswith("⚠") or "skipped" in msg or "not found" in msg.lower():
            self._log(msg, "warn")
        elif "SUPPLEMENT" in msg:
            self._log(msg, "head")
        else:
            self._log(msg)

    def _clear_log(self):
        self.log_box.config(state="normal")
        self.log_box.delete("1.0", "end")
        self.log_box.config(state="disabled")

    def _on_run(self):
        docx   = self.shared["var_docx"].get().strip()
        excel  = self.shared["var_excel"].get().strip()
        out    = self.var_out.get().strip()
        review = self.var_review.get().strip() or None
        pdf    = self.var_pdf.get().strip() or None

        errors = []
        if not docx or not os.path.isfile(docx):
            errors.append("Prospectus introuvable.")
        if not excel or not os.path.isfile(excel):
            errors.append("Fichier Excel introuvable.")
        if not out:
            errors.append("Chemin de sortie normale manquant.")
        if errors:
            messagebox.showerror("Erreur", "\n".join(errors))
            return

        self.btn_run.config(state="disabled")
        self.progress.start(12)
        self._clear_log()
        self._log("─" * 60)
        self._log(f"Prospectus    : {docx}")
        self._log(f"Mapping       : {excel}")
        self._log(f"Sortie normale: {out}")
        if review:
            self._log(f"Sortie review : {review}")
        if pdf:
            self._log(f"Sortie PDF    : {pdf}")
        self._log("─" * 60)

        dry = self.var_dry_run.get()

        def _task():
            try:
                result = run(
                    docx, excel,
                    out    + ".DRY_RUN_IGNORE" if dry else out,
                    (review + ".DRY_RUN_IGNORE" if dry else review) if review else None,
                    None if dry else pdf,
                    log=self._smart_log,
                )
                if dry:
                    self._log("\n[DRY RUN] Aucun fichier sauvegardé.", "warn")
                elif self.var_open_after.get():
                    if os.path.isfile(out):
                        self.after(500, lambda: _open_file(out))
                    if review and os.path.isfile(review):
                        self.after(800, lambda: _open_file(review))
                    if pdf and os.path.isfile(pdf):
                        self.after(1100, lambda: _open_file(pdf))

                n_warn = len(result["warnings"])
                self._log(
                    f"\n{'✓' if not n_warn else '⚠'} Terminé — "
                    f"{result['insertions']} insertion(s), {n_warn} avertissement(s).",
                    "warn" if n_warn else "ok",
                )
            except Exception as exc:
                self._log(f"\n✗ Erreur : {exc}", "err")
                self.after(0, lambda: messagebox.showerror("Erreur", str(exc)))
            finally:
                self.after(0, lambda: (self.progress.stop(),
                                       self.btn_run.config(state="normal")))

        threading.Thread(target=_task, daemon=True).start()


# ─────────────────────────────────────────────────────────────────────────────
# Content editor widget (embedded in TabClauses)
# ─────────────────────────────────────────────────────────────────────────────

class ContentEditor(ttk.LabelFrame):
    """
    Dynamic editor that adapts to the clause type:
      texte       → single multi-line text area
      liste       → list of bullet entries (add / delete rows)
      sous_titres → list of (subtitle, body-text) row pairs (add / delete)
    """

    def __init__(self, parent, **kwargs):
        super().__init__(parent, text="Contenu", padding=6, **kwargs)
        self._type = None
        self._rows  = []   # list of row frames + StringVar(s)
        self._text_widget = None
        self._items_frame = None
        self.set_type("texte")

    # ── Public API ────────────────────────────────────────────────────────────

    def set_type(self, type_):
        if type_ == self._type:
            return
        self._type = type_
        self._rows = []
        for w in self.winfo_children():
            w.destroy()

        if type_ == "texte":
            self._build_texte()
        elif type_ == "liste":
            self._build_liste()
        elif type_ == "sous_titres":
            self._build_sous_titres()

    def get_items(self):
        """Return list of dicts matching the mapping.py content format."""
        if self._type == "texte":
            return [{"texte": self._text_widget.get("1.0", "end-1c").strip(),
                     "sous_texte": ""}]
        items = []
        for row_data in self._rows:
            if self._type == "liste":
                t = row_data[0].get().strip()
                if t:
                    items.append({"texte": t, "sous_texte": ""})
            else:  # sous_titres
                t  = row_data[0].get().strip()
                st = row_data[1].get().strip()
                if t:
                    items.append({"texte": t, "sous_texte": st})
        return items

    def set_items(self, items):
        """Load content items into the editor."""
        # Rebuild to clear existing rows
        self.set_type(self._type)
        if not items:
            return
        if self._type == "texte":
            self._text_widget.delete("1.0", "end")
            self._text_widget.insert("end", items[0].get("texte", ""))
        elif self._type == "liste":
            for item in items:
                self._add_row(item.get("texte", ""))
        else:  # sous_titres
            for item in items:
                self._add_row(item.get("texte", ""), item.get("sous_texte", ""))

    # ── Type-specific builders ────────────────────────────────────────────────

    def _build_texte(self):
        self._text_widget = scrolledtext.ScrolledText(
            self, height=5, wrap="word", font=("TkDefaultFont", 9))
        self._text_widget.pack(fill="both", expand=True)

    def _build_liste(self):
        self._items_frame = ttk.Frame(self)
        self._items_frame.pack(fill="x")
        ttk.Button(self, text="+ Ajouter une puce",
                   command=lambda: self._add_row()).pack(anchor="w", pady=(4, 0))

    def _build_sous_titres(self):
        hdr = ttk.Frame(self)
        hdr.pack(fill="x")
        ttk.Label(hdr, text="Sous-titre", width=20).pack(side="left")
        ttk.Label(hdr, text="Texte").pack(side="left", padx=(4, 0))
        self._items_frame = ttk.Frame(self)
        self._items_frame.pack(fill="x")
        ttk.Button(self, text="+ Ajouter un sous-titre",
                   command=lambda: self._add_row()).pack(anchor="w", pady=(4, 0))

    # ── Row management ────────────────────────────────────────────────────────

    def _add_row(self, text="", sous_text=""):
        row_frame = ttk.Frame(self._items_frame)
        row_frame.pack(fill="x", pady=1)

        if self._type == "liste":
            ttk.Label(row_frame, text="•", width=2).pack(side="left")
            var = tk.StringVar(value=text)
            ttk.Entry(row_frame, textvariable=var).pack(
                side="left", fill="x", expand=True, padx=(2, 4))
            ttk.Button(row_frame, text="✕", width=2,
                       command=lambda f=row_frame, v=var: self._del_row(f, v)
                       ).pack(side="left")
            self._rows.append((var,))

        else:  # sous_titres
            var_t  = tk.StringVar(value=text)
            var_st = tk.StringVar(value=sous_text)
            ttk.Entry(row_frame, textvariable=var_t, width=20).pack(
                side="left", padx=(0, 4))
            ttk.Entry(row_frame, textvariable=var_st).pack(
                side="left", fill="x", expand=True, padx=(0, 4))
            ttk.Button(row_frame, text="✕", width=2,
                       command=lambda f=row_frame, vt=var_t, vs=var_st:
                           self._del_row(f, vt, vs)
                       ).pack(side="left")
            self._rows.append((var_t, var_st))

    def _del_row(self, frame, *vars_):
        frame.destroy()
        self._rows = [r for r in self._rows if r[0] is not vars_[0]]


# ─────────────────────────────────────────────────────────────────────────────
# Tab 2 — Gestion des clauses
# ─────────────────────────────────────────────────────────────────────────────

class TabClauses(ttk.Frame):
    def __init__(self, parent, shared):
        super().__init__(parent)
        self.shared = shared
        self._editing_id    = None
        self._clause_content = {}   # {clause_id: [items]}
        self._build()
        self._load_from_excel()

    # ── Layout ────────────────────────────────────────────────────────────────

    def _build(self):
        self.columnconfigure(0, weight=3)
        self.columnconfigure(1, weight=2)
        self.rowconfigure(0, weight=1)

        # ── LEFT: Treeview ────────────────────────────────────────────────────
        frm_tree = ttk.LabelFrame(self, text="Clauses définies", padding=6)
        frm_tree.grid(row=0, column=0, sticky="nsew", padx=(12, 4), pady=10)
        frm_tree.rowconfigure(0, weight=1)
        frm_tree.columnconfigure(0, weight=1)

        cols = ("id", "titre", "ancre", "position", "type")
        self.tree = ttk.Treeview(frm_tree, columns=cols,
                                 show="headings", selectmode="browse")
        self.tree.heading("id",       text="ID",       anchor="w")
        self.tree.heading("titre",    text="Titre",    anchor="w")
        self.tree.heading("ancre",    text="Ancre",    anchor="w")
        self.tree.heading("position", text="Position", anchor="w")
        self.tree.heading("type",     text="Type",     anchor="w")
        self.tree.column("id",       width=50,  stretch=False)
        self.tree.column("titre",    width=140)
        self.tree.column("ancre",    width=140)
        self.tree.column("position", width=90,  stretch=False)
        self.tree.column("type",     width=80,  stretch=False)

        sb = ttk.Scrollbar(frm_tree, orient="vertical", command=self.tree.yview)
        self.tree.config(yscrollcommand=sb.set)
        self.tree.grid(row=0, column=0, sticky="nsew")
        sb.grid(row=0, column=1, sticky="ns")
        self.tree.bind("<<TreeviewSelect>>", self._on_tree_select)

        frm_tree_btn = ttk.Frame(frm_tree)
        frm_tree_btn.grid(row=1, column=0, columnspan=2, sticky="ew", pady=(6, 0))
        ttk.Button(frm_tree_btn, text="Nouvelle clause",
                   command=self._new_clause).pack(side="left")
        self.btn_delete = ttk.Button(frm_tree_btn, text="Supprimer",
                                     command=self._delete_clause, state="disabled")
        self.btn_delete.pack(side="left", padx=(6, 0))

        # ── RIGHT: Form ───────────────────────────────────────────────────────
        frm_right = ttk.Frame(self)
        frm_right.grid(row=0, column=1, sticky="nsew", padx=(4, 12), pady=10)
        frm_right.columnconfigure(0, weight=1)
        frm_right.rowconfigure(1, weight=1)

        frm_meta = ttk.LabelFrame(frm_right, text="Définir / Modifier", padding=8)
        frm_meta.grid(row=0, column=0, sticky="ew")
        frm_meta.columnconfigure(1, weight=1)

        self.var_id       = tk.StringVar()
        self.var_titre    = tk.StringVar()
        self.var_ancre    = tk.StringVar()
        self.var_position = tk.StringVar(value="apres_titre")
        self.var_type     = tk.StringVar(value="texte")

        ttk.Label(frm_meta, text="ID").grid(row=0, column=0, sticky="w", pady=3)
        ttk.Entry(frm_meta, textvariable=self.var_id).grid(
            row=0, column=1, columnspan=2, sticky="ew", padx=(6, 0))

        ttk.Label(frm_meta, text="Titre").grid(row=1, column=0, sticky="w", pady=3)
        ttk.Entry(frm_meta, textvariable=self.var_titre).grid(
            row=1, column=1, columnspan=2, sticky="ew", padx=(6, 0))

        ttk.Label(frm_meta, text="Ancre").grid(row=2, column=0, sticky="w", pady=3)
        ttk.Entry(frm_meta, textvariable=self.var_ancre).grid(
            row=2, column=1, sticky="ew", padx=(6, 0))
        ttk.Button(frm_meta, text="⌕", width=3,
                   command=self._open_anchor_picker).grid(row=2, column=2, padx=(4, 0))

        ttk.Label(frm_meta, text="Position").grid(row=3, column=0, sticky="w", pady=3)
        ttk.Combobox(frm_meta, textvariable=self.var_position,
                     values=["apres_titre", "apres_section"],
                     state="readonly", width=16).grid(
                         row=3, column=1, sticky="w", padx=(6, 0))

        ttk.Label(frm_meta, text="Type").grid(row=4, column=0, sticky="w", pady=3)
        cb_type = ttk.Combobox(frm_meta, textvariable=self.var_type,
                               values=["texte", "liste", "sous_titres"],
                               state="readonly", width=16)
        cb_type.grid(row=4, column=1, sticky="w", padx=(6, 0))
        cb_type.bind("<<ComboboxSelected>>",
                     lambda _: self.content_editor.set_type(self.var_type.get()))

        # Anchor preview
        self.lbl_preview = ttk.Label(frm_meta, text="", wraplength=200,
                                     foreground="#777",
                                     font=("TkDefaultFont", 8, "italic"))
        self.lbl_preview.grid(row=5, column=0, columnspan=3, sticky="w",
                              padx=(0, 0), pady=(2, 0))
        self.var_ancre.trace_add("write",
            lambda *_: self.lbl_preview.config(text=self.var_ancre.get()[:100]))

        # Content editor (dynamic)
        self.content_editor = ContentEditor(frm_right)
        self.content_editor.grid(row=1, column=0, sticky="nsew", pady=(8, 0))

        # Save / Reset buttons
        frm_btn = ttk.Frame(frm_right)
        frm_btn.grid(row=2, column=0, sticky="ew", pady=(8, 0))
        ttk.Button(frm_btn, text="Enregistrer",
                   command=self._save_clause).pack(side="right")
        ttk.Button(frm_btn, text="Réinitialiser",
                   command=self._reset_form).pack(side="right", padx=(0, 6))

        # ── Bottom: Excel actions ─────────────────────────────────────────────
        frm_bottom = ttk.Frame(self)
        frm_bottom.grid(row=1, column=0, columnspan=2,
                        sticky="ew", padx=12, pady=(0, 10))
        self.lbl_status = ttk.Label(frm_bottom, text="", foreground="green")
        self.lbl_status.pack(side="left")
        ttk.Button(frm_bottom, text="💾  Sauvegarder dans l'Excel",
                   command=self._save_to_excel).pack(side="right")
        ttk.Button(frm_bottom, text="↺  Recharger depuis l'Excel",
                   command=self._load_from_excel).pack(side="right", padx=(0, 8))

    # ── Tree helpers ──────────────────────────────────────────────────────────

    def _tree_rows(self):
        return [self.tree.item(iid)["values"] for iid in self.tree.get_children()]

    def _refresh_tree(self, rows):
        self.tree.delete(*self.tree.get_children())
        for row in rows:
            self.tree.insert("", "end", values=row)

    def _on_tree_select(self, _=None):
        sel = self.tree.selection()
        if not sel:
            self.btn_delete.config(state="disabled")
            return
        self.btn_delete.config(state="normal")
        v = self.tree.item(sel[0])["values"]
        self._editing_id = str(v[0])
        self.var_id.set(v[0])
        self.var_titre.set(v[1])
        self.var_ancre.set(v[2])
        self.var_position.set(v[3] if len(v) > 3 else "apres_titre")
        typ = str(v[4]) if len(v) > 4 else "texte"
        self.var_type.set(typ)
        self.content_editor.set_type(typ)
        self.content_editor.set_items(
            self._clause_content.get(str(v[0]), []))

    # ── Form actions ──────────────────────────────────────────────────────────

    def _new_clause(self):
        self.tree.selection_remove(self.tree.selection())
        self._reset_form()

    def _reset_form(self):
        self._editing_id = None
        self.var_id.set("")
        self.var_titre.set("")
        self.var_ancre.set("")
        self.var_position.set("apres_titre")
        self.var_type.set("texte")
        self.content_editor.set_type("texte")
        self.content_editor.set_items([])

    def _save_clause(self):
        cid      = self.var_id.get().strip()
        titre    = self.var_titre.get().strip()
        ancre    = self.var_ancre.get().strip()
        position = self.var_position.get()
        typ      = self.var_type.get()

        if not cid or not titre or not ancre:
            messagebox.showwarning("Champs manquants",
                                   "ID, Titre et Ancre sont obligatoires.")
            return

        rows = self._tree_rows()
        new_row = [cid, titre, ancre, position, typ]

        if self._editing_id is None:
            if any(str(r[0]) == cid for r in rows):
                messagebox.showerror("ID en double", f"L'ID « {cid} » existe déjà.")
                return
            rows.append(new_row)
        else:
            rows = [new_row if str(r[0]) == self._editing_id else r for r in rows]

        self._clause_content[cid] = self.content_editor.get_items()
        self._refresh_tree(rows)
        self._reset_form()
        self._set_status("Clause enregistrée (non sauvegardée dans l'Excel)")

    def _delete_clause(self):
        sel = self.tree.selection()
        if not sel:
            return
        cid = str(self.tree.item(sel[0])["values"][0])
        if not messagebox.askyesno("Confirmer", f"Supprimer la clause « {cid} » ?"):
            return
        self.tree.delete(sel[0])
        self._clause_content.pop(cid, None)
        self._reset_form()
        self._set_status(f"Clause {cid} supprimée")

    # ── Anchor picker ─────────────────────────────────────────────────────────

    def _open_anchor_picker(self):
        docx = self.shared["var_docx"].get().strip()
        if not docx or not os.path.isfile(docx):
            messagebox.showwarning("Prospectus manquant",
                "Sélectionnez d'abord un prospectus dans l'onglet Traitement.")
            return
        AnchorPickerDialog(self.winfo_toplevel(), docx, self.var_ancre)

    # ── Excel I/O ─────────────────────────────────────────────────────────────

    def _load_from_excel(self):
        excel = self.shared["var_excel"].get().strip()
        if not excel or not os.path.isfile(excel):
            return
        try:
            wb   = openpyxl.load_workbook(excel)
            rows = []
            for r in wb["clauses"].iter_rows(min_row=2, values_only=True):
                if not r[0]:
                    continue
                pos = str(r[3]).strip().lower() if len(r) > 3 and r[3] else "apres_titre"
                typ = str(r[4]).strip().lower() if len(r) > 4 and r[4] else "texte"
                rows.append([str(r[0]), str(r[1] or ""), str(r[2] or ""), pos, typ])

            # Load content
            self._clause_content = {}
            if "contenu" in wb.sheetnames:
                for r in wb["contenu"].iter_rows(min_row=2, values_only=True):
                    if not r[0]:
                        continue
                    cid  = str(r[0]).strip()
                    texte     = str(r[2]).strip() if len(r) > 2 and r[2] else ""
                    sous_texte = str(r[3]).strip() if len(r) > 3 and r[3] else ""
                    self._clause_content.setdefault(cid, []).append(
                        {"texte": texte, "sous_texte": sous_texte})

            self._refresh_tree(rows)
            self._set_status(f"{len(rows)} clause(s) chargée(s)")
        except Exception as exc:
            messagebox.showerror("Erreur", f"Impossible de lire l'Excel :\n{exc}")

    def _save_to_excel(self):
        excel = self.shared["var_excel"].get().strip()
        if not excel:
            messagebox.showerror("Erreur", "Aucun fichier Excel sélectionné.")
            return
        rows = self._tree_rows()
        if not rows:
            messagebox.showwarning("Vide", "Aucune clause à sauvegarder.")
            return
        try:
            wb = openpyxl.load_workbook(excel)

            # ── clauses sheet ─────────────────────────────────────────────────
            ws1 = wb["clauses"]
            for row in ws1.iter_rows(min_row=2):
                for cell in row:
                    cell.value = None
            for i, r in enumerate(rows, start=2):
                for j, val in enumerate(r, start=1):
                    ws1.cell(row=i, column=j, value=val)

            # ── contenu sheet ─────────────────────────────────────────────────
            if "contenu" not in wb.sheetnames:
                wb.create_sheet("contenu")
            ws2 = wb["contenu"]
            # Ensure header
            if ws2.max_row < 1 or ws2.cell(1, 1).value != "ClauseID":
                ws2.delete_rows(1, ws2.max_row)
                ws2.append(["ClauseID", "Ordre", "Texte", "Sous_texte"])
            else:
                ws2.delete_rows(2, ws2.max_row)

            row_idx = 2
            for r in rows:
                cid   = str(r[0])
                items = self._clause_content.get(cid, [])
                for ordre, item in enumerate(items, start=1):
                    ws2.cell(row=row_idx, column=1, value=cid)
                    ws2.cell(row=row_idx, column=2, value=ordre)
                    ws2.cell(row=row_idx, column=3, value=item.get("texte", ""))
                    ws2.cell(row=row_idx, column=4, value=item.get("sous_texte", ""))
                    row_idx += 1

            wb.save(excel)
            self._set_status(f"✓ Sauvegardé dans {os.path.basename(excel)}")
        except Exception as exc:
            messagebox.showerror("Erreur", f"Impossible d'écrire dans l'Excel :\n{exc}")

    def _set_status(self, msg):
        self.lbl_status.config(text=msg)
        self.after(5000, lambda: self.lbl_status.config(text=""))


# ─────────────────────────────────────────────────────────────────────────────
# Main App
# ─────────────────────────────────────────────────────────────────────────────

class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("CPR Invest – Clause Inserter")
        self.resizable(True, True)
        self.minsize(780, 600)

        # Shared variables (docx + excel paths, used by both tabs)
        base = os.path.dirname(os.path.abspath(__file__))
        shared = {
            "var_docx":  tk.StringVar(value=os.path.join(base, "data", "prospectus.docx")),
            "var_excel": tk.StringVar(value=os.path.join(base, "data", "mapping.xlsx")),
        }

        notebook = ttk.Notebook(self)
        notebook.pack(fill="both", expand=True, padx=6, pady=6)

        tab1 = TabTraitement(notebook, shared)
        tab2 = TabClauses(notebook, shared)

        notebook.add(tab1, text="  Traitement  ")
        notebook.add(tab2, text="  Gestion des clauses  ")

        # Default output path
        tab1.var_out.set(os.path.join(base, "data", "prospectus_updated.docx"))
        tab1.var_review.set(os.path.join(base, "data", "prospectus_review.docx"))
        tab1.var_pdf.set(os.path.join(base, "data", "prospectus_updated.pdf"))

        # Reload clauses when switching to tab 2
        notebook.bind("<<NotebookTabChanged>>", lambda e: (
            tab2._load_from_excel()
            if notebook.index(notebook.select()) == 1 else None
        ))


# ─────────────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    app = App()
    app.mainloop()
