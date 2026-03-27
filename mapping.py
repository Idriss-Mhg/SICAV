import openpyxl

VALID_TYPES     = ("texte", "liste", "sous_titres")
VALID_POSITIONS = ("apres_titre", "apres_section")


def load_mapping(excel_path):
    """
    Returns:
        clauses : dict { clause_id -> {title, anchor, position, type, content} }
        mapping : dict { compartment_name -> [clause_id, ...] }

    Excel structure
    ───────────────
    Sheet "clauses"  : ClauseID | ClauseTitre | InsererApres | Position | Type
    Sheet "contenu"  : ClauseID | Ordre | Texte | Sous_texte
    Sheet "mapping"  : Compartiment | CL01 | CL02 | ...  (X = active)
    """
    wb = openpyxl.load_workbook(excel_path)

    # ── Sheet "clauses" ───────────────────────────────────────────────────────
    ws_clauses = wb["clauses"]
    clauses = {}
    for row in ws_clauses.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        clause_id  = str(row[0]).strip()
        title      = str(row[1]).strip() if row[1] else ""
        anchor     = str(row[2]).strip() if row[2] else ""
        pos        = str(row[3]).strip().lower() if len(row) > 3 and row[3] else "apres_titre"
        typ        = str(row[4]).strip().lower() if len(row) > 4 and row[4] else "texte"
        exact_pos  = str(row[5]).strip() if len(row) > 5 and row[5] else ""

        if pos not in VALID_POSITIONS:
            pos = "apres_titre"
        if typ not in VALID_TYPES:
            typ = "texte"

        clauses[clause_id] = {
            "title":     title,
            "anchor":    anchor,
            "position":  pos,
            "type":      typ,
            "content":   [],   # filled below from "contenu" sheet
            "exact_pos": exact_pos,  # optional: overrides automatic position detection
        }

    # ── Sheet "contenu" (optional) ────────────────────────────────────────────
    if "contenu" in wb.sheetnames:
        ws_contenu = wb["contenu"]
        for row in ws_contenu.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue
            cid       = str(row[0]).strip()
            texte     = str(row[2]).strip() if len(row) > 2 and row[2] else ""
            sous_texte = str(row[3]).strip() if len(row) > 3 and row[3] else ""
            if cid in clauses:
                clauses[cid]["content"].append({
                    "texte":      texte,
                    "sous_texte": sous_texte,
                })

    # ── Sheet "mapping" ───────────────────────────────────────────────────────
    ws_mapping = wb["mapping"]
    header     = [cell.value for cell in ws_mapping[1]]
    clause_ids = [str(v).strip() for v in header[1:] if v]

    mapping = {}
    for row in ws_mapping.iter_rows(min_row=2, values_only=True):
        comp_name = row[0]
        if not comp_name:
            continue
        active = []
        for j, val in enumerate(row[1: len(clause_ids) + 1]):
            if val and str(val).strip().upper() in ("X", "YES", "OUI", "1", "TRUE"):
                active.append(clause_ids[j])
        mapping[str(comp_name).strip()] = active

    return clauses, mapping
