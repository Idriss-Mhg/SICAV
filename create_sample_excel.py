"""
Run once to generate data/mapping.xlsx with sample data.
"""
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

OUTPUT = "data/mapping.xlsx"
HDR    = Font(bold=True)
FILL   = PatternFill("solid", fgColor="D9E1F2")


def _style_header(ws, col_widths):
    for cell in ws[1]:
        cell.font = HDR
        cell.fill = FILL
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width


def main():
    wb = openpyxl.Workbook()

    # ── Sheet "clauses" ───────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "clauses"
    ws1.append(["ClauseID", "ClauseTitre", "InsererApres", "Position", "Type"])

    sample_clauses = [
        ["CL01", "Sustainable Finance Disclosure",
         "Risk Management:",     "apres_section", "texte"],
        ["CL02", "Stewardship Policy",
         "Conflicts of interest", "apres_section", "liste"],
        ["CL03", "German Investment Tax Act",
         "Reference Currency:",  "apres_titre",   "sous_titres"],
    ]
    for row in sample_clauses:
        ws1.append(row)

    _style_header(ws1, {"A": 10, "B": 38, "C": 42, "D": 16, "E": 14})

    # ── Sheet "contenu" ───────────────────────────────────────────────────────
    ws2 = wb.create_sheet("contenu")
    ws2.append(["ClauseID", "Ordre", "Texte", "Sous_texte"])

    sample_content = [
        # CL01 — texte
        ["CL01", 1,
         "The Management Company ensures compliance with Regulation (EU) 2019/2088 "
         "(SFDR). Pre-contractual and periodic disclosures are made available to investors "
         "in accordance with the applicable regulatory requirements.",
         ""],
        # CL02 — liste (puces)
        ["CL02", 1, "Monitor and engage with investee companies on ESG matters.", ""],
        ["CL02", 2, "Exercise voting rights in accordance with the Stewardship Policy.", ""],
        ["CL02", 3, "Disclose engagement and voting outcomes in the annual report.", ""],
        # CL03 — sous_titres
        ["CL03", 1,
         "Equity fund status",
         "At least 51% of the Compartment's assets consist of equity participations "
         "within the meaning of the German Investment Tax Act (InvStG)."],
        ["CL03", 2,
         "Mixed fund status",
         "At least 25% of the Compartment's assets consist of equity participations "
         "within the meaning of the German Investment Tax Act (InvStG)."],
    ]
    for row in sample_content:
        ws2.append(row)

    _style_header(ws2, {"A": 10, "B": 8, "C": 55, "D": 55})

    # ── Sheet "mapping" ───────────────────────────────────────────────────────
    ws3 = wb.create_sheet("mapping")
    clause_ids = [r[0] for r in sample_clauses]
    ws3.append(["Compartiment"] + clause_ids)

    sample_mapping = [
        ["CPR Invest \u2013 Silver Age",  "X", "X", ""],
        ["CPR Invest \u2013 Reactive",    "X", "X", "X"],
    ]
    for row in sample_mapping:
        ws3.append(row)

    _style_header(ws3, {"A": 32, "B": 8, "C": 8, "D": 8})
    for row in ws3.iter_rows(min_row=2):
        for cell in row[1:]:
            cell.alignment = Alignment(horizontal="center")

    os.makedirs(os.path.dirname(OUTPUT), exist_ok=True)
    wb.save(OUTPUT)
    print(f"Created: {OUTPUT}")


if __name__ == "__main__":
    main()
